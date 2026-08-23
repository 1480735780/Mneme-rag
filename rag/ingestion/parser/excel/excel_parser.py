# -*- coding: utf-8 -*-
"""
Excel 文档解析器（openpyxl 能力等价替代 Apache POI，对应 Java ExcelDocumentParser）

单元格规范化交给 ExcelTableNormalizer：合并单元格展开填充、多行表头展平拼接、
超链接内联为 [text](url)、公式回退缓存值或公式字符串。

每个非隐藏 sheet 产出 0 或 2 个 Block：
    - HeadingBlock(level=1, text=sheet 名)：sheet 名走 HeadingBlock 而不是自建上下文字段，
      H1 的顶级重置语义正好是 sheet 之间的关系，交给 HeadingHandler 维护 outlinePath 后，
      sheet 名自然落到向量文本前缀。
    - TableBlock(headers, rows)

对应 ragent 源码：
    - com.nageoffer.ai.ragent.core.parser.excel.ExcelDocumentParser
"""
from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Optional, Set

from common.exception.business import ServiceException
from rag.ingestion.parser.base import DocumentParser, ParseProfile, ParserType
from rag.ingestion.parser.excel.hyperlink_resolver import wrap
from rag.ingestion.parser.excel.table_normalizer import STRIKETHROUGH_WRAP, normalize
from rag.ingestion.parser.excel.value_formatter import format_cell, is_strikethrough
from rag.ingestion.parser.model import HeadingBlock, ParsedDocument, Provenance, TableBlock

_OPT_SOURCE_FILE = "sourceFile"
_OPT_HEADER_ROWS = "headerRows"
_DEFAULT_HEADER_ROWS = 1


class ExcelDocumentParser(DocumentParser):
    """Excel 文档解析器（openpyxl）：xlsx → 每 sheet 一个 HeadingBlock + TableBlock"""

    @property
    def parser_type(self) -> str:
        return ParserType.EXCEL_POI.value

    def supported_mime_types(self) -> Dict[ParseProfile, Set[str]]:
        # 快速档承担全部表格类，含 Tika 的两个 Office 家族别名
        return {
            ParseProfile.FAST: {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
                "application/x-tika-msoffice",
                "application/x-tika-ooxml",
            }
        }

    def parse_structured(
        self,
        content: bytes,
        mime_type: Optional[str] = None,
        options: Optional[Dict[str, object]] = None,
    ) -> ParsedDocument:
        if not content:
            return ParsedDocument.of([])

        source_file = self._extract_string(options)
        header_rows = self._extract_int(options)

        from openpyxl import load_workbook

        try:
            # 公式需两份：data_only=False 拿公式字符串，data_only=True 拿缓存值（openpyxl 无实时求值）
            wb_formula = load_workbook(BytesIO(content), data_only=False, read_only=False)
            wb_cached = load_workbook(BytesIO(content), data_only=True, read_only=False)
        except Exception as e:  # noqa: BLE001
            raise ServiceException(f"Excel 解析失败: {e}") from e

        blocks: List[object] = []
        total_sheets = len(wb_formula.sheetnames)
        try:
            for name in wb_formula.sheetnames:
                ws = wb_formula[name]
                if ws.sheet_state in ("hidden", "veryHidden"):
                    continue
                ws_cached = wb_cached[name] if name in wb_cached.sheetnames else None

                def render_cell(cell, _cached=ws_cached) -> str:
                    """单元格渲染：格式化 + 超链接内联 + 删除线包裹"""
                    value = cell.value
                    formula = isinstance(value, str) and value.startswith("=")
                    cached = None
                    if formula and _cached is not None:
                        cached = _cached[cell.coordinate].value
                    text = format_cell(value, formula, cached)
                    text = wrap(text, cell)
                    if text and is_strikethrough(cell):
                        text = STRIKETHROUGH_WRAP + text + STRIKETHROUGH_WRAP
                    return text

                table = normalize(ws, render_cell, header_rows)
                if table.is_empty():
                    continue
                prov = Provenance.of_excel_cell(source_file, name)
                blocks.append(HeadingBlock(prov, 1, name))
                blocks.append(TableBlock(prov, table.headers, table.rows))
        finally:
            wb_formula.close()
            wb_cached.close()

        return ParsedDocument.of(
            blocks,
            {
                "parser": self.parser_type,
                "mimeType": mime_type or "",
                "totalSheets": total_sheets,
                # 只数表格：每个 sheet 还额外产一个承载 sheet 名的 HeadingBlock
                "parsedTables": sum(1 for b in blocks if isinstance(b, TableBlock)),
                "headerRows": header_rows,
            },
        )

    # ------------------------------------------------------------------ #

    @staticmethod
    def _extract_string(options: Optional[Dict[str, object]]) -> str:
        if not options:
            return ""
        v = options.get(_OPT_SOURCE_FILE)
        return "" if v is None else str(v)

    @staticmethod
    def _extract_int(options: Optional[Dict[str, object]]) -> int:
        if not options:
            return _DEFAULT_HEADER_ROWS
        v = options.get(_OPT_HEADER_ROWS)
        if v is None:
            return _DEFAULT_HEADER_ROWS
        if isinstance(v, int):
            return v
        try:
            return int(str(v))
        except (TypeError, ValueError):
            return _DEFAULT_HEADER_ROWS
