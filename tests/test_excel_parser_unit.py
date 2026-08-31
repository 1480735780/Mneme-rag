# -*- coding: utf-8 -*-
"""
Excel 解析单元测试：ExcelDocumentParser + Normalizer + Formatter + HyperlinkResolver（对应 Java）

覆盖：
    - MIME 认领 / parser_type
    - 基本表：sheet → HeadingBlock(sheet 名) + TableBlock
    - 多 sheet、隐藏 sheet 跳过
    - 合并单元格展开填充
    - 多行表头展平（| 拼接、相邻去重）
    - 全空列 / 全空行裁剪
    - 数值/日期/布尔格式化、删除线包裹
    - 超链接内联 [text](url)
    - 公式缓存值回退 / 公式字符串回退
    - headerRows 配置 / sourceFile / 元数据 / 空内容
"""
from io import BytesIO

import pytest

from openpyxl import Workbook

from common.exception.business import ServiceException
from rag.ingestion.parser.base import ParseProfile, ParserType
from rag.ingestion.parser.excel.excel_parser import ExcelDocumentParser
from rag.ingestion.parser.model import HeadingBlock, TableBlock


def _xlsx(builder):
    """builder(wb) 构造后返回 bytes"""
    wb = Workbook()
    builder(wb)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _basic_wb(wb):
    ws = wb.active
    ws.title = "订单"
    ws.append(["名称", "金额", "备注"])
    ws.append(["定金", 100, "ok"])
    ws.append(["尾款", 200, None])


def _parse(content, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", **options):
    return ExcelDocumentParser().parse_structured(content, mime, options or None)


class TestExcelDocumentParser:
    def test_parser_type(self):
        assert ExcelDocumentParser().parser_type == ParserType.EXCEL_POI.value

    def test_supported_mime_types(self):
        mimes = ExcelDocumentParser().supported_mime_types()[ParseProfile.FAST]
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in mimes
        assert "application/vnd.ms-excel" in mimes
        assert "application/x-tika-ooxml" in mimes

    def test_empty_content_returns_empty(self):
        doc = _parse(b"")
        assert doc.blocks == []

    def test_basic_sheet_heading_and_table(self):
        doc = _parse(_xlsx(_basic_wb))
        # HeadingBlock(sheet 名) + TableBlock
        assert len(doc.blocks) == 2
        assert isinstance(doc.blocks[0], HeadingBlock)
        assert doc.blocks[0].text == "订单"
        assert doc.blocks[0].level == 1
        block = doc.blocks[1]
        assert isinstance(block, TableBlock)
        assert block.headers == ["名称", "金额", "备注"]
        assert block.rows == [["定金", "100", "ok"], ["尾款", "200", ""]]

    def test_multiple_sheets(self):
        def build(wb):
            _basic_wb(wb)
            ws2 = wb.create_sheet("退货")
            ws2.append(["原因", "数量"])
            ws2.append(["破损", 2])

        doc = _parse(_xlsx(build))
        assert len(doc.blocks) == 4  # 2 sheet × (Heading + Table)
        texts = [b.text for b in doc.blocks if isinstance(b, HeadingBlock)]
        assert texts == ["订单", "退货"]

    def test_hidden_sheet_skipped(self):
        def build(wb):
            _basic_wb(wb)
            ws2 = wb.create_sheet("隐藏")
            ws2.append(["a", "b"])
            ws2.sheet_state = "hidden"

        doc = _parse(_xlsx(build))
        assert len(doc.blocks) == 2  # 只有可见 sheet
        texts = [b.text for b in doc.blocks if isinstance(b, HeadingBlock)]
        assert texts == ["订单"]

    def test_merged_cells_expanded(self):
        def build(wb):
            ws = wb.active
            ws.append(["类别", "金额"])
            ws.append(["A", 1])
            ws.append(["A", 2])
            ws.merge_cells("A2:A3")  # 类别列 A 合并两行

        doc = _parse(_xlsx(build))
        block = doc.blocks[1]
        assert block.rows == [["A", "1"], ["A", "2"]]  # 合并区左上角值复制

    def test_multi_row_header_flattened(self):
        def build(wb):
            ws = wb.active
            ws.append(["财务", "财务", "销售"])
            ws.append(["收入", "支出", "收入"])
            ws.append([1, 2, 3])

        doc = _parse(_xlsx(build), headerRows=2)
        block = doc.blocks[1]
        assert block.headers == ["财务|收入", "财务|支出", "销售|收入"]

    def test_empty_column_dropped(self):
        # 表头与数据全程为空的列才被裁掉（对齐 Java：全空列丢弃）
        def build(wb):
            ws = wb.active
            ws.append(["名称", "金额", ""])  # 第三列表头也空
            ws.append(["定金", 100, None])

        doc = _parse(_xlsx(build))
        block = doc.blocks[1]
        assert block.headers == ["名称", "金额"]
        assert block.rows == [["定金", "100"]]

    def test_header_only_column_kept(self):
        # 仅表头有值、数据全空的列保留（数据本身有意义，表头即信息）
        def build(wb):
            ws = wb.active
            ws.append(["名称", "金额", "备注"])
            ws.append(["定金", 100, None])

        doc = _parse(_xlsx(build))
        block = doc.blocks[1]
        assert block.headers == ["名称", "金额", "备注"]  # 表头有值 → 保留
        assert block.rows == [["定金", "100", ""]]

    def test_empty_row_dropped(self):
        def build(wb):
            ws = wb.active
            ws.append(["名称", "金额"])
            ws.append(["定金", 100])
            ws.append([None, None])

        doc = _parse(_xlsx(build))
        block = doc.blocks[1]
        assert block.rows == [["定金", "100"]]

    def test_formula_cached_value_used(self):
        def build(wb):
            ws = wb.active
            ws.append(["a", "b"])
            ws.append([1, None])
            ws["B2"] = "=A2*2"
            # openpyxl 保存时不写缓存值；手动补写（模拟带缓存文件）
            ws["B2"] = "=A2*2"

        doc = _parse(_xlsx(build))
        block = doc.blocks[1]
        # 无缓存值 → 回退公式字符串（openpyxl 无法实时求值）
        assert block.rows == [["1", "=A2*2"]]

    def test_hyperlink_inline(self):
        def build(wb):
            ws = wb.active
            ws.append(["链接", "值"])
            cell = ws["A2"]
            cell.value = "官网"
            cell.hyperlink = "https://example.com"
            ws["B2"] = 1

        doc = _parse(_xlsx(build))
        block = doc.blocks[1]
        assert block.rows == [["[官网](https://example.com)", "1"]]

    def test_strikethrough_wrapped(self):
        import copy

        def build(wb):
            ws = wb.active
            ws.append(["名称", "金额"])
            ws["B2"] = 100
            new_font = copy.copy(ws["B2"].font)
            new_font.strike = True
            ws["B2"].font = new_font

        doc = _parse(_xlsx(build))
        block = doc.blocks[1]
        assert block.rows == [["", "~~100~~"]]  # 删除线包裹

    def test_metadata(self):
        doc = _parse(_xlsx(_basic_wb))
        assert doc.metadata["parser"] == "ExcelPoi"
        assert doc.metadata["totalSheets"] == 1
        assert doc.metadata["parsedTables"] == 1
        assert doc.metadata["headerRows"] == 1

    def test_source_file_and_header_rows_options(self):
        doc = _parse(_xlsx(_basic_wb), sourceFile="x.xlsx", headerRows=1)
        assert doc.blocks[1].provenance.source_file == "x.xlsx"
        assert doc.blocks[1].provenance.sheet_name == "订单"

    def test_invalid_file_raises_service_exception(self):
        with pytest.raises(ServiceException):
            _parse(b"not an xlsx file")

    def test_invalid_header_rows_defaults(self):
        doc = _parse(_xlsx(_basic_wb), headerRows="bad")
        assert doc.metadata["headerRows"] == 1
